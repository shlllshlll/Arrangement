# -*- coding: utf-8 -*-
# @Author:  SHLLL
# @email:   shlll7347@gmail.com
# @Date:    2018-09-15 00:01:57
# @License: MIT LICENSE
# @Last Modified by:   Mr.Shi
# @Last Modified time: 2018-09-21 14:56:03

from openpyxl import load_workbook
from .. import json


class PersonData(object):
    """Person data initializer."""

    def __init__(self, data_struct):
        self._data_struct = data_struct
        self._base_path = data_struct['base_path']

    def working(self, type):
        """人员数据处理函数

        选择从xlsx文件读取数据或者从Json读取数据或者
        将数据存入Json文件中

        Args:
            type: 1:从xlsx中读取数据并存入Json并返回数据
                  2:从Json中读取数据并返回
                  3:将数据存入Json (default: {1})
        """
        if type == 1:
            self._read_from_xlsx()
        elif type == 2:
            self._read_from_json()
        elif type == 3:
            self._write_to_json()

    def _read_from_json(self):
        data_struct = self._data_struct
        # 输出的数据
        data_struct['datedata'] = json.read_json(
            self._base_path + 'json/DateData.json')
        data_struct['peopledata'] = json.read_json(
            self._base_path + 'json/PeopleData.json')
        data_struct['people_id'] = json.read_json(
            self._base_path + 'json/PeopleID.json')

    def _write_to_json(self):
        data_struct = self._data_struct
        # 将数据写入Json
        json.write_json(data_struct['peopledata'],
                        self._base_path + 'json/PeopleData.json')
        json.write_json(data_struct['datedata'],
                        self._base_path + 'json/DateData.json')
        json.write_json(data_struct['people_id'],
                        self._base_path + 'json/PeopleID.json')

    def _read_from_xlsx(self):
        """从Xlsx文件中读取数据函数"""
        data_struct = self._data_struct
        xlsx_filename = self._base_path + 'xlsx/轮转信息表.xlsx'

        # 获取Excel工作表
        wb = load_workbook(filename=xlsx_filename)
        ws = wb.active

        # 获取人员与ID对应关系
        people_id = []

        # ------按照人名获取排班信息-------
        # Get the row number of the sheet
        rows = ws.max_row
        date = [x.value for x in ws['1'][1:]]
        people_data = []
        for i in range(2, rows + 1):
            # 记录人员及其对应ID号
            people_id.append({'name': ws[str(i)][0].value, 'id': i - 1})
            person_data = {}
            person_data['name'] = ws[str(i)][0].value
            person_data['id'] = i - 1
            temp = [x.value == '骨' for x in ws[str(i)][1:]]
            person_data['times'] = temp.count(True)
            # 将值班日期提取出来存入数组
            month_data = []
            for i, val in enumerate(temp):
                if val is True:
                    month_data.append(date[i])
            person_data['month'] = month_data
            people_data.append(person_data)
            person_data['history'] = []
            person_data['avliable'] = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        json.write_json(people_data, self._base_path + 'json/PeopleData.json')

        # ------记录人员与ID号对应关系------
        json.write_json(people_id, self._base_path + 'json/PeopleID.json')

        # ------按照日期获取排班名单------
        date_data = []
        cols = ws.max_column
        for i in range(2, cols + 1):
            month_data = {}
            month_data['date'] = ws.cell(1, i).value
            people = []
            people_mon_id = []
            temp = [x.value == '骨' for x in ws[str(i)][1:]]
            for cells in ws.iter_cols(i, i):
                for cell in cells:
                    if cell.value == '骨':
                        people.append(ws.cell(cell.row, 1).value)
                        people_mon_id.append(cell.row - 1)
            month_data['num'] = len(people_mon_id)
            month_data['people'] = people
            month_data['people_id'] = people_mon_id
            date_data.append(month_data)
        json.write_json(date_data, self._base_path + 'json/DateData.json')

        # 输出的数据
        data_struct['datedata'] = date_data
        data_struct['peopledata'] = people_data
        data_struct['people_id'] = people_id
