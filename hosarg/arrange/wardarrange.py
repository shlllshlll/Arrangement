# -*- coding: utf-8 -*-
# @Author:  SHLLL
# @email:   shlll7347@gmail.com
# @Date:    2018-09-20 14:48:44
# @License: MIT LICENSE
# @Last Modified by:   Mr.Shi
# @Last Modified time: 2018-09-21 16:15:38

import math
import datetime
import calendar
from random import shuffle
from openpyxl import load_workbook, Workbook
from hosarg.utils.datapreprocess import PersonData


class WardArrange(object):

    def __init__(self, data, date):
        self._data = data
        # 获取科室分组名单
        departs = self._read_depart()
        # 获取四个病房的人员名单
        wards = self._slice_people(departs)
        name_list = self._arrange_ward(wards, date)
        self._save_to_xslx(name_list)

    def _read_depart(self, file_name='科室分组.xlsx'):
        file_path = self._data['base_path'] + 'xlsx/' + file_name
        wb = load_workbook(file_path)
        ws = wb.active
        departs = []
        for i, col in enumerate(ws.iter_cols(1, 6)):
            depart = []
            for cell in col[2:]:
                if cell.value is not None:
                    # 读取排班人员
                    val = cell.value.replace('*', '').\
                        replace('#', '').replace('^', '')
                    depart.append(val)
                    self._fresh_person_history(val, i + 1)
            departs.append(depart)
        # 回存数据
        PersonData(self._data).working(3)
        return departs

    def _fresh_person_history(self, name, depart_id):
        people_data = self._data['peopledata']
        name_list = [i['name'] for i in people_data]
        try:
            person_idx = name_list.index(name)
        except ValueError:
            pass
        else:
            person = people_data[person_idx]
            try:
                depart_idx = person['history'].index(depart_id)
            except ValueError:
                person['history'].append(depart_id)

    def _arrange_ward(self, wards, date):
        # 首先对时间进行处理
        date = int(date)
        # 分离获取年月
        year = 2000 + int(date / 100)
        month = date % 100
        # 获取本月有多少天
        month_range = calendar.monthrange(year, month)[1]

        print('-----------')
        people_in_wards = []
        # 取出每一个病房中的人员名单
        for ward in wards:
            people_in_ward = []
            people_dict = []
            # 将人员名单转换为一个dict
            for person in ward:
                people_dict.append(
                    {'name': person, 'history': [], 'weekday': []})
            print('++++++++')
            print('/'.join([i['name'] for i in people_dict]))
            # 遍历本月的每一天
            for day in range(1, month_range + 1):
                cur_date = datetime.date(year, month, day)
                # 获取当前是周几
                wkd = cur_date.isoweekday()
                is_weekdat = 1 if wkd < 6 else 0
                person = self._get_a_person(day, is_weekdat, people_dict)
                people_in_ward.append(person)
            print(people_in_ward)
            people_in_wards.append(people_in_ward)

        return people_in_wards

    def _get_a_person(self, day, weekday, people):
        # 获取存储的数据，没有则创建数据

        result = None
        for idx, person in enumerate(people):

            # 如果尚未值过班
            if not person['history']:
                result = person['name']
                person['history'].append(day)
                person['weekday'].append(weekday)
                item = people.pop(idx)
                people.append(item)
                break
            else:
                lst_day = person['history'][-1]
                lst_wkd = person['weekday'][-1]
                if day - lst_day > 5 and weekday - lst_wkd != 0:
                    result = person['name']
                    person['history'].append(day)
                    person['weekday'].append(weekday)
                    item = people.pop(idx)
                    people.append(item)
                    break

        # 如果找了一圈都没有满足条件的指定为第一个
        if not result:
            item = people.pop(0)
            people.append(item)
            result = item['name']

        return result

    def _save_to_xslx(self, data, file_name='病房排班表.xlsx'):
        wb = Workbook()
        ws = wb.active

        # 首先添加列名称
        col_name = ['日期', '骨1', '骨2', '骨3', '骨4']
        ws.append(col_name)

        # 接着添加人员信息
        for day in range(0, len(data[0])):
            ws.append([day + 1, data[0][day], data[1][day],
                       data[2][day], data[3][day]])
        wb.save(self._data['base_path'] + 'result/' + file_name)

    def _slice_people(self, departs):
        wards = []

        # 预处理数据
        shuffle(departs[0])
        shuffle(departs[1])
        shuffle(departs[2])
        shuffle(departs[3])
        shuffle(departs[4])
        shuffle(departs[5])
        luzhui = departs[0]
        jingzhui = departs[1]
        guanjie = departs[2]
        chuangshang = departs[3]
        jizhui = departs[4]
        yaozhui = departs[5]

        # 首先分配骨1病房
        ward1 = []
        ward1.extend(jingzhui[:2])
        ward1.extend(guanjie[:math.floor(len(guanjie) / 2)])
        ward1.extend(jizhui[:math.floor(len(jizhui) / 2)])
        wards.append(ward1)

        # 分配骨2病房
        ward2 = []
        ward2.extend(jizhui[math.floor(len(jizhui) / 2):])
        ward2.extend(yaozhui)
        wards.append(ward2)

        # 分配骨3病房
        ward3 = []
        ward3.extend(guanjie[math.floor(len(guanjie) / 2):])
        ward3.extend(chuangshang)
        wards.append(ward3)

        # 分配骨4病房
        ward4 = []
        ward4.extend(luzhui)
        ward4.extend(jingzhui[2:])
        wards.append(ward4)

        print(ward1)
        print(ward2)
        print(ward3)
        print(ward4)

        return wards
