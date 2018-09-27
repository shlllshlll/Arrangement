# -*- coding: utf-8 -*-
# @Author:  SHLLL
# @email:   shlll7347@gmail.com
# @Date:    2018-09-14 20:38:57
# @License: MIT LICENSE
# @Last Modified by:   Mr.Shi
# @Last Modified time: 2018-09-21 14:46:56
import copy
import random
from prettytable import PrettyTable
from openpyxl import Workbook
from hosarg.utils.datapreprocess import PersonData


class DepartArrange(object):
    """Department arrangement class"""

    def __init__(self, data, date):
        # 获取基础数据
        self._data = data
        self._people_data = data['peopledata']
        date = str(date)
        depart_rule = data['departrule']
        date_datas = data['datedata']

        # 获取科室数量
        departs_num = len(depart_rule)

        # 创建一个科室容器用于盛放人员安排
        departs_container = []
        for i in range(0, departs_num):
            # depart_num = depart_rule[i]['max']
            # depart_container = [0 for m in range(depart_num)]
            depart_container = []
            departs_container.append(depart_container)

        # 读取科室规则
        departs_rule_max = [i['max'] for i in depart_rule]
        departs_rule_radio = [i['radio'] for i in depart_rule]
        departs_ruls = {'max': departs_rule_max, 'radio': departs_rule_radio}

        # 准备人员数据
        # 获取月份列表
        months_data = [i['date'] for i in date_datas]
        month_idx = months_data.index(date)
        date_data = date_datas[month_idx]
        # 获取本月排班人员的ID号和人数
        people_ids = date_data['people_id']
        # people_num = date_data['num']
        # 获取本月有排班人员的数据列表
        people_cur_data = self._find_people(people_ids)

        # 依次读取人员数据放入科室容器中
        self._greedy_alg(people_cur_data,
                         departs_ruls,
                         departs_container)
        self._pretty_print(departs_container)
        self._save_xlsx(departs_container)

    def _find_people(self, id_list):
        """根据指定id列表查找对应的人员数据函数

        本函数接受一个人员id列表，遍历列表，
        返回对应id号人员的所有信息

        Args:
            id_list: 人员id号列表

        Returns:
            返回找打的人员数据
            返回list列表
        """
        people_ids = [i['id'] for i in self._people_data]

        people_data = []
        for id in id_list:
            idx = people_ids.index(id)
            people_data.append(self._people_data[idx])
        return people_data

    def _greedy_alg(self, data, rules, container):
        """核心算法函数

        执行排班算法的函数

        Args:
            data: 输入的人员数据
            rules: 输入的科室规则
            container: 输入的科室排班容器
        """
        rules_max = rules['max']
        rules_radio = rules['radio']
        cur_radio = self._cal_radio(container)
        # 记录安排失败的人员名单
        person_failed = []

        for person in data:
            # 人员安排成功标志位
            person_succeed_flag = False
            # 打乱科室选择顺序，从而使得结果多样化
            person_avli = copy.copy(person['avliable'])
            random.shuffle(person_avli)
            for i, depart in enumerate(person_avli):
                # 使用贪心算法，即只要满足约束条件就安排其进科室
                # 首先检查人数是否达到科室容量
                if(len(container[depart - 1]) >= rules_max[depart - 1]):
                    continue
                # 其次检查人员比例是否合理
                if cur_radio[depart - 1] > rules_radio[depart - 1]:
                    continue

                # 将当前的人员放入容器中
                container[depart - 1].append(person['id'])
                # 更新人员数据中的排班历史课和可排班数据
                person['history'].append(depart)
                person['avliable'].pop(person['avliable'].index(depart))
                # 更新当前radio
                cur_radio = self._cal_radio(container)
                person_succeed_flag = True
                break
            # 如果当前循环次数
            if person_succeed_flag is False:
                person_failed.append(person['name'])
        self._person_failed = person_failed
        PersonData(self._data).working(3)

    def _cal_radio(self, container):
        """计算当前科室人员比例函数

        Args:
            container: 科室排班容器
        """
        length = 0
        length_list = []
        for item in container:
            length += len(item)
            length_list.append(len(item))

        if length == 0:
            return [0. for x in length_list]
        else:
            return [x / length for x in length_list]

    def _container2str(self, container):
        depart_id = self._data['departid']
        people_id = self._data['people_id']
        table = []
        for i, item in enumerate(container):
            column_name = depart_id[i + 1]
            people_name = [people_id[i - 1]['name'] for i in item]
            table.append([column_name] + people_name)
        return table

    def _pretty_print(self, container):
        depart_id = self._data['departid']
        people_id = self._data['people_id']
        max_len = max([len(i) for i in container])
        x = PrettyTable()
        for i, item in enumerate(container):
            column_name = depart_id[i + 1]
            people_name = [people_id[i - 1]['name'] for i in item]
            x.add_column(column_name, people_name +
                         ['' for i in range(max_len - len(item))])
        print(x)
        print('--------以下人员排班失败---------')
        print(','.join(self._person_failed))

    def _save_xlsx(self, container):
        table = self._container2str(container)
        base_path = self._data['base_path']
        wb = Workbook()
        ws = wb.active
        ws.title = '排班信息表'

        for m, col in enumerate(table):
            for n, item in enumerate(col):
                ws.cell(n + 1, m + 1).value = item
        ws.cell(1, m + 2).value = '排班失败名单'
        for i, item in enumerate(self._person_failed):
            ws.cell(i + 2, m + 2).value = item
        wb.save(base_path + 'result/' + '科室排班.xlsx')
